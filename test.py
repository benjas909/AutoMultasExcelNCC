import openpyxl, re
import tkinter as tk
from openpyxl.styles import Font, Alignment, PatternFill
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from copy import copy
import os
import itertools

import openpyxl.styles

DEFAULT_DIR = "."
ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


# Activa la opción de importar tickets al seleccionar la casilla correspondiente.
def activator():

    if check.get() == 1:
        importButton.config(state="normal")
    else:
        importButton.config(state="disabled")


def copySheet(sourceSheet, targetSheet):
    """
    Copia por completo una hoja.

    args:
        sourceSheet(object): Hoja de origen a copiar.
        targetSheet(object): Hoja de destino a pegar.
    
    Función extraída de StackOverflow https://stackoverflow.com/a/68800310
    """
    copyCells(sourceSheet, targetSheet)
    copySheetAttributes(sourceSheet, targetSheet)



def copyCells(sourceSheet, targetSheet):
    """
    Copia contenidos y estilos de cada celda de una hoja

    args:
        sourceSheet(object): Hoja de origen de las celdas.
        targetSheet(object): Hoja de destino de las celdas.

    Función extraída de StackOverflow https://stackoverflow.com/a/68800310
    """    
    for (row, col), sourceCell in sourceSheet._cells.items():
        targetCell = targetSheet.cell(column=col, row=row)

        targetCell._value = sourceCell._value
        targetCell.data_type = sourceCell.data_type

        if sourceCell.has_style:
            targetCell.font = copy(sourceCell.font)
            targetCell.border = copy(sourceCell.border)
            targetCell.fill = copy(sourceCell.fill)
            targetCell.number_format = copy(sourceCell.number_format)
            targetCell.protection = copy(sourceCell.protection)
            targetCell.alignment = copy(sourceCell.alignment)

        if sourceCell.hyperlink:
            targetCell._hyperlink = copy(sourceCell.hyperlink)

        if sourceCell.comment:
            targetCell.comment = copy(sourceCell.comment)
            

def copySheetAttributes(sourceSheet, targetSheet):
    """
    Copia atributos de una hoja.

    args:
        sourceSheet(object): Hoja de origen a copiar.
        targetSheet(object): Hoja de destino a pegar.

    Función extraída de StackOverflow: https://stackoverflow.com/a/68800310
    """
    targetSheet.sheet_format = copy(sourceSheet.sheet_format)
    targetSheet.sheet_properties = copy(sourceSheet.sheet_properties)
    targetSheet.merged_cells = copy(sourceSheet.merged_cells)
    targetSheet.page_margins = copy(sourceSheet.page_margins)
    targetSheet.freeze_panes = copy(sourceSheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(sourceSheet.row_dimensions)):
        targetSheet.row_dimensions[rn] = copy(sourceSheet.row_dimensions[rn])

    if sourceSheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        targetSheet.sheet_format.defaultColWidth = copy(sourceSheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in sourceSheet.column_dimensions.items():
        targetSheet.column_dimensions[key].min = copy(sourceSheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        targetSheet.column_dimensions[key].max = copy(sourceSheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        targetSheet.column_dimensions[key].width = copy(sourceSheet.column_dimensions[key].width) # set width for every column
        targetSheet.column_dimensions[key].hidden = copy(sourceSheet.column_dimensions[key].hidden)


def XLSXHandling(ticketsFilename, inFilename, outFilename):
    """
    Realiza la manipulación de las planillas xlsx.

    args:
        ticketsFilename(str): Nombre del archivo xlsx que contiene la información de tickets.
        inFilename(str): Nombre del archivo xlsx con información de multas.
        outFilename(str): Nombre de archivo de salida.
    """
    workbook = openpyxl.load_workbook(filename=inFilename)
    if (ticketsFilename != ""):
        targetSheet = workbook.create_sheet("Tickets")
        tickWorkbook = openpyxl.load_workbook(filename=ticketsFilename)
        sourceSheet = tickWorkbook["Sheet 1"]
        copySheet(sourceSheet, targetSheet)

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    # Abrir archivo excel de entrada
    ticketsSheet = workbook["Tickets"]
    tickMax = ticketsSheet.max_row # Número de fila del último ticket registrado en la hoja de tickets
    print(tickMax)
    sheet = workbook["Item 12"] # Selección de hoja
    lastTicket = sheet.max_row - 3 # Número de fila del último ticket de la hoja original

    contents = []

    # Guarda contenidos de hoja 
    for row in sheet.iter_rows():
        contents.append(row)

    cyan = openpyxl.styles.colors.Color(rgb="0000b0F0")
    headerFont = Font(name="Calibri", size=9, bold=True, color="00FFFFFF")
    headerFill = PatternFill(start_color=cyan, end_color=cyan, fill_type="solid")
    centerAlign = Alignment(horizontal="center", vertical="center")

    # Header titles
    sheet["J1"] = "Apertura"
    sheet["K1"] = "Resuelto"
    sheet["L1"] = "Cerrado"
    sheet["M1"] = "Mejor Fecha"
    sheet["N1"] = "Tiempo Resolución"
    sheet["O1"] = "Tiempo Cierre"
    sheet["P1"] = "Mejor Tiempo"
    sheet["Q1"] = "Solicitud Apertura ATM"
    sheet["R1"] = "Disponibilidad Apertura ATM"
    sheet["S1"] = "Tiempo Descuento"
    sheet["T1"] = "Tiempo Indisponibilidad GTD"
    sheet["U1"] = "Responsable"

    # Header styling 
    for letter in ALPHABET:
        if letter > 'I' and letter < 'V':
            sheet[f"{letter}1"].fill = headerFill
            
        sheet[f"{letter}1"].font = headerFont
        sheet[f"{letter}1"].alignment = centerAlign

    i = 1
    addedCells = 0
    for row in contents:

        # Deja de leer al pasar del último ATM
        if (row[0].value is None):
            break

        sheet[f"C{i}"] = str(row[2].value)
        numList = re.findall(r"\d{6}", row[2].value)

        # Si ATM tiene por lo menos un ticket
        if len(numList) > 0:
            sheet[f"C{i}"] = "2024 " + numList[0] # Reescribe número de ticket con formato correcto
            sheet[f"J{i}"] = f'=IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},22,FALSE)), "No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},22,FALSE))'
            sheet[f"K{i}"] = f'=IF(ISBLANK(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)),"Vacío",IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)),"No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)))'
            sheet[f"L{i}"] = f'=IF(ISBLANK(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)),"Vacío",IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)),"No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)))'
            sheet[f"N{i}"] = f'=IF(ISERR(K{i}-J{i}), "No Disponible", K{i}-J{i})'
            sheet[f"O{i}"] = f'=IF(ISERR(L{i}-J{i}), "No Disponible", L{i}-J{i})'
            sheet[f"P{i}"] = f'=IF(ISTEXT(N{i}),IF(ISTEXT(O{i}),"No disponible",O{i}),IF(ISTEXT(O{i}),N{i},MIN(M{i}, N{i},O{i})))'
            sheet[f"S{i}"] = f'=R{i}-Q{i}'
            sheet[f"T{i}"] = f'=IF(ISERR(P{i}-S{i}), "No Disponible", P{i}-S{i})'
            sheet[f"U{i}"] = f'=IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},15,FALSE)), "No Disponible", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},15,FALSE))'

            # Si ATM tiene más de un ticket
            if len(numList) > 1:
                # Guarda datos de ATM actual
                ATM = row[0].value
                Comuna = row[1].value

                # Recorre lista de tickets, desde el segundo ticket
                for item in numList[1:]:
                    newNum = f"2024 {item}"
                    sheet.insert_rows(idx= i + 1)
                    addedCells += 1
                    i += 1
                    sheet[f"A{i}"] = ATM
                    sheet[f"B{i}"] = Comuna
                    sheet[f"C{i}"] = newNum
                    sheet[f"J{i}"] = f'=IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},22,FALSE)), "No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},22,FALSE))'
                    sheet[f"K{i}"] = f'=IF(ISBLANK(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)),"Vacío",IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)),"No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},23,FALSE)))'
                    sheet[f"L{i}"] = f'=IF(ISBLANK(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)),"Vacío",IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)),"No Encontrado", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},24,FALSE)))'
                    sheet[f"N{i}"] = f'=IF(ISERR(K{i}-J{i}), "No Disponible", K{i}-J{i})'
                    sheet[f"O{i}"] = f'=IF(ISERR(L{i}-J{i}), "No Disponible", L{i}-J{i})'
                    sheet[f"P{i}"] = f'=IF(ISTEXT(N{i}),IF(ISTEXT(O{i}),"No disponible",O{i}),IF(ISTEXT(O{i}),N{i},MIN(M{i}, N{i},O{i})))'
                    sheet[f"S{i}"] = f'=R{i}-Q{i}'
                    sheet[f"T{i}"] = f'=IF(ISERR(P{i}-S{i}), "No Disponible", P{i}-S{i})'
                    sheet[f"U{i}"] = f'=IF(ISNA(VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},15,FALSE)), "No Disponible", VLOOKUP(C{i},Tickets!$A$2:$X${tickMax},15,FALSE))'


        # ATM sin número de ticket
        elif(len(numList) == 0 and i != 1):
            sheet[f"J{i}"] = "No Disponible"
            sheet[f"K{i}"] = "No Disponible"
            sheet[f"L{i}"] = "No Disponible"
            sheet[f"N{i}"] = "No Disponible"
            sheet[f"O{i}"] = "No Disponible"

        i += 1

    sheet[f"I{lastTicket + addedCells + 3}"] = f'=SUM(I2:I{lastTicket + addedCells})'

    # Junta las columnas que tienen formato de fecha
    dateCells1 = sheet[ f'J1:M{lastTicket + addedCells + 3}' ]
    dateCells2 = sheet[ f'Q1:R{lastTicket + addedCells + 3}' ]
    dateCellsJoin = [dateCells1, dateCells2]

    # Junta las columnas que tienen formato de hora
    timeCells1 = sheet[ f"N1:P{lastTicket + addedCells + 3}" ]
    timeCells2 = sheet[ f"S1:U{lastTicket + addedCells + 3}" ]
    timeCellsJoin = [timeCells1, timeCells2]

    # Aplica estilos a la hoja
    for r in sheet[ f"A2:V{lastTicket + addedCells + 3}" ]:
        for cell in r:
            cell.font = Font(name = "Calibri", size = 9)
            cell.alignment = Alignment(horizontal = "center", vertical = "center")

    # Aplica formato de fecha a columnas correspondientes
    for r in itertools.chain(*dateCellsJoin):
        for cell in r:
            cell.number_format = "dd/mm/yyyy h:mm"

    # Aplica formato de hora a columnas correspondientes
    for r in itertools.chain(*timeCellsJoin):
        for cell in r:
            cell.number_format = "[h]:mm:ss"

    # Aplica ancho de columna
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 20
    sheet.column_dimensions["L"].width = 20
    sheet.column_dimensions["M"].width = 20
    sheet.column_dimensions["N"].width = 15
    sheet.column_dimensions["O"].width = 15
    sheet.column_dimensions["P"].width = 20
    sheet.column_dimensions["Q"].width = 25
    sheet.column_dimensions["R"].width = 25
    sheet.column_dimensions["S"].width = 15
    sheet.column_dimensions["T"].width = 25
    sheet.column_dimensions["U"].width = 15

    workbook.save(filename=outFilename)
    showinfo(title="Listo", message="Archivo guardado")


def selectInputFile(inType):
    """
    Ventana emergente para selección de archivos.

    args:
        inType(str): Tipo de archivo a abrir ("multas" o "tickets). 
    """

    global inputFilename, ticketsFilename
    filetypes = (
        ('Excel spreadsheets', '*.xlsx'),
        ('All files', '*.*')
    )

    if (inType == "multas"):
        inputFilename = fd.askopenfilename(
            title = "Abrir archivo",
            initialdir=DEFAULT_DIR,
            filetypes=filetypes
        )
        if inputFilename:
            showinfo(title="Archivo seleccionado", message=inputFilename)
            if hasattr(window, "label_infile"):
                window.label_infile.config(text="Archivo de multas: " + inputFilename)
        else:
            showinfo(title="Error", message="No se ha seleccionado ningún archivo")

    elif (inType == "tickets"):
        ticketsFilename = fd.askopenfilename(
            title = "Abrir archivo de tickets",
            initialdir = DEFAULT_DIR,
            filetypes = filetypes
        )
        if ticketsFilename:
            showinfo(title = "Archivo de tickets seleccionado", message = ticketsFilename)
            if hasattr(window, "label_tickfile"):
                window.label_tickfile.config(text="Archivo de tickets: " + ticketsFilename)
            else:
                showinfo(title = "Error", message = "No se ha seleccionado un archivo de tickets")


def saveAs() :
    """
    Ventana emergente para selección de ubicación de destino y nombre de archivo de salida.
    """
    global inputFilename, outputFilename

    # Si no se ha seleccionado un archivo de entrada, se muestra error y se cierra el diálogo
    if not inputFilename:
        showinfo(title="Error", message="No se ha seleccionado un archivo de multas.")
        return

    baseName = os.path.splitext(os.path.basename(inputFilename))[0]
    print(baseName)

    defaultName = f"{baseName}_processed.xlsx"

    filetypes = (
        ("Excel Spreadsheets", "*.xlsx"), 
        ("All Files", "*.*")
    )

    outputFilename = fd.asksaveasfilename(title = "Guardar como", defaultextension = ".xlsx",
     initialfile = defaultName, initialdir = DEFAULT_DIR, filetypes = filetypes)
    
    if outputFilename:
        showinfo(title="", message=outputFilename)
        if hasattr(window, "label_outfile"):
            window.label_outfile.config(text="Nombre de salida: " + outputFilename)
        XLSXHandling(ticketsFilename, inputFilename, outputFilename)
    else:
        showinfo(title="Error", message="No se ha seleccionado un nombre de salida")


############# INTERFAZ #############
# Ventana principal
window = tk.Tk()
window.resizable(False, False)
window.title("AutoMultas")
window.geometry("400x400")

ticketsFilename = ""
inputFilename = ""
outputFilename = ""

# Frame de importación de tickets
impFrame = tk.Frame(master=window, width=450, height=150, bd=4, relief=tk.RAISED)
impFrame.pack(pady=5)

# Checkbox de activación de botón de importe de tickets
check = tk.IntVar()
importCheck = ttk.Checkbutton(impFrame, text="Importar tickets", variable=check, onvalue=1, offvalue=0, command=activator)
importCheck.pack(pady=5)

# Label de nombre de archivo de tickets y botón de importe  
window.label_tickfile = tk.Label(impFrame, text="", wraplength=300)
window.label_tickfile.pack()
importButton = ttk.Button(master=impFrame, text="Abrir archivo de tickets", command= lambda: selectInputFile("tickets"), state="disabled")
importButton.pack(padx=10, pady=5)

# Frame de apertura de archivo de multas
openFrame = tk.Frame(master=window, width=450, height=100, bd=4, relief=tk.RAISED)
openFrame.pack(pady=5)

# Label de nombre de archivo y botón de abrir archivo
window.label_infile = tk.Label(openFrame, text="No se ha seleccionado un archivo de multas", wraplength=300)
window.label_infile.pack()
openButton = ttk.Button(master=openFrame, text="Abrir un archivo de multas", command= lambda: selectInputFile("multas"))
openButton.pack(pady=10)

# Frame de guardado de archivo
saveFrame = tk.Frame(master=window, width=450, height=100, bd=4, relief=tk.RAISED)
saveFrame.pack()

# Label de nombre de archivo y botón de "Guardar Como..."
window.label_outfile = tk.Label(saveFrame, text="No se ha seleccionado un nombre de archivo de salida", wraplength=300)
window.label_outfile.pack()
saveButton = ttk.Button(master=saveFrame, text="Guardar archivo", command=saveAs)
saveButton.pack(pady=10)


window.mainloop()

############# INTERFAZ #############


